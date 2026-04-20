"""Sample Python file for anvil-py locator tests.

The shape echoes the author's openpyxl report-automation usage:
from-import, decorated function, async function, classes with
methods, a nested helper, and a module-level constant.  Locators
must find all of these cleanly."""

from __future__ import annotations

import os
import sys as system
from pathlib import Path
from openpyxl import Workbook, load_workbook
from typing import (
    Iterable,
    Optional,
)


MAX_ROWS = 1000


def _open(path):
    return load_workbook(path)


def plain_func(a, b):
    return a + b


async def fetch_value(url):
    return url


@staticmethod
def static_helper(x):
    return x * 2


@classmethod
@property
def multi_decorated(cls):
    return cls


class ReportWriter:
    """Write a report workbook."""

    def __init__(self, path):
        self.path = path

    def write(self, rows):
        wb = Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(self.path)

    @staticmethod
    def default_name() -> str:
        return "report.xlsx"

    async def write_async(self, rows):
        self.write(rows)


@dataclass
class Config:
    name: str
    limit: int = 100

    def describe(self):
        return f"{self.name}: {self.limit}"


def outer():
    def inner(a):
        return a + 1
    return inner
